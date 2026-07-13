#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""BOQ Price Finder — ค้นราคาวัสดุจาก 4 แหล่ง + custom URL, เก็บเป็น Price Library,
export Excel และเติมราคากลับเข้าไฟล์ BOQ

รัน:  python3 server.py   แล้วเปิด http://127.0.0.1:5544
"""
import datetime
import json
import os
import re
import threading
import urllib.parse
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from bs4 import BeautifulSoup
from flask import Flask, jsonify, request, send_file, send_from_directory

BASE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE, 'data')
LIB_FILE = os.path.join(DATA_DIR, 'price_library.json')
OBEC_FILE = os.path.join(DATA_DIR, 'obec2568.json')
os.makedirs(DATA_DIR, exist_ok=True)

UA = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                    'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125 Safari/537.36'}
TIMEOUT = 15
app = Flask(__name__)
_lib_lock = threading.Lock()

THAI_MARKS = re.compile(u'[ัิ-ฺ็-๎\\s\\.\\-/]')
MENU_TEXT_RE = re.compile(
    r'(Search Product|Category|Home|About\s*US|Product|Download|Price List|Contact\s*US|Shop Online|'
    r'ตู้คอนซูเมอร์|ตู้โหลดเซ็นเตอร์|มิเตอร์ไฟฟ้า|©\s*2016)',
    re.I
)
PRICE_UNIT_RE = re.compile(r'([\d,]+(?:\.\d+)?)\s*/\s*([^\s,|]+)', re.I)
GENERIC_NAME_RE = re.compile(r'^(ชนิด|รายการ|ขนาด|ราคา|model|code|type|spec|description)$', re.I)


# ---------------------------------------------------------------- common parser engine
def norm(s):
    """normalize สำหรับค้นแบบทนตัวสะกด/วรรณยุกต์เพี้ยน (ข้อมูลจาก PDF)"""
    return THAI_MARKS.sub('', (s or '').lower())


def today():
    return datetime.date.today().isoformat()


def fetch_url(url, method="GET", params=None, data=None, headers=None, timeout=TIMEOUT):
    h = dict(UA)
    if headers:
        h.update(headers)
    if method.upper() == "POST":
        r = requests.post(url, params=params, data=data, headers=h, timeout=timeout)
    else:
        r = requests.get(url, params=params, headers=h, timeout=timeout)
    r.raise_for_status()
    r.encoding = r.apparent_encoding or r.encoding or 'utf-8'
    return r


def parse_price(text):
    m = re.search(r'([\d,]+(?:\.\d+)?)', text or '')
    if not m:
        return None
    try:
        return float(m.group(1).replace(',', ''))
    except ValueError:
        return None


def expand_query_aliases(q):
    raw = compact_text(q)
    parts = [p for p in re.split(r'\s+', raw) if p]
    aliases = parts[:]
    upper = raw.upper()
    if 'IEC' in upper:
        aliases += ['IEC', '60227', '60227 IEC', '60227 IEC 01', 'THW', 'YAZAKI', 'ยาซากิ', 'สายไฟ']
    if 'THW' in upper:
        aliases += ['THW', 'IEC', '60227', '60227 IEC 01', 'YAZAKI', 'ยาซากิ', 'สายไฟ']
    if 'YAZAKI' in upper or 'ยาซากิ' in raw:
        aliases += ['YAZAKI', 'ยาซากิ', 'IEC', '60227', '60227 IEC 01', 'THW', 'สายไฟ']
    if 'สายไฟ' in raw:
        aliases += ['สายไฟ', 'IEC', 'THW', '60227', 'YAZAKI', 'ยาซากิ']
    seen, out = set(), []
    for alias in aliases:
        key = alias.lower()
        if alias and key not in seen:
            seen.add(key)
            out.append(alias)
    return out


def make_result(name, price, unit="", labor=None, source="", url="", parser_name="", matched_terms=None, confidence=0.0):
    return {
        'name': compact_text(name),
        'price': price,
        'unit': compact_text(unit),
        'labor': labor,
        'source': compact_text(source),
        'url': url or '',
        'fetched_at': today(),
        'matched_terms': matched_terms or [],
        'parser_name': parser_name,
        'confidence': float(confidence or 0.0),
    }


def make_source_status(status, result_count=0, message=""):
    return {'status': status, 'result_count': int(result_count or 0), 'message': message or ''}


def parse_price_unit(text):
    text = compact_text(text)
    patterns = [
        r'([\d,]+(?:\.\d+)?)\s*/\s*([^\s,|]+)',
        r'ราคา\s*([\d,]+(?:\.\d+)?)\s*บาท(?:\s*/\s*([^\s,]+))?',
        r'([\d,]+(?:\.\d+)?)\s*บาท(?:\s*/\s*([^\s,]+))?',
    ]
    for pat in patterns:
        m = re.search(pat, text, re.I)
        if m:
            return parse_price(m.group(1)), (m.group(2) if len(m.groups()) > 1 and m.group(2) else '')
    return None, ''


def query_tokens(q):
    return [t for t in re.split(r'\s+', (q or '').strip().lower()) if t]


def is_relevant_result(item, q):
    tokens = query_tokens(q)
    if not tokens:
        return True
    hay = ' '.join(str(item.get(k, '') or '').lower() for k in ('name', 'url', 'source'))
    # Short English searches like IEC are often ignored by shop search endpoints.
    # Keep only rows that visibly contain the searched token.
    if all(re.fullmatch(r'[a-z0-9]{2,4}', t) for t in tokens):
        return all(t in hay for t in tokens)
    thai_tokens = [norm(t) for t in tokens if re.search(r'[ก-๙]', t)]
    if thai_tokens:
        name_norm = norm(item.get('name', ''))
        return any(t in name_norm for t in thai_tokens) or any(t in hay for t in tokens if re.search(r'[a-z0-9]', t))
    return all(t in hay for t in tokens)


def source_origin(url):
    parsed = urllib.parse.urlparse(url if re.match(r'^https?://', url, re.I) else 'https://' + url)
    return '%s://%s' % (parsed.scheme, parsed.netloc)


def build_custom_search_url(url, q):
    url = url.strip()
    if not re.match(r'^https?://', url, re.I):
        url = 'https://' + url
    if '{q}' in url:
        return url.replace('{q}', urllib.parse.quote(q))
    parsed = urllib.parse.urlparse(url)
    params = urllib.parse.parse_qs(parsed.query, keep_blank_values=True)
    key = next((k for k in ('q', 's', 'search', 'keyword', 'keywords') if k in params), 's')
    params[key] = [q]
    query = urllib.parse.urlencode(params, doseq=True)
    return urllib.parse.urlunparse(parsed._replace(query=query))


def absolutize_url(href, base):
    return urllib.parse.urljoin(base, href or '')


def compact_text(text):
    return re.sub(r'\s+', ' ', text or '').strip()


def is_truncated_text(text):
    return '...' in (text or '') or '…' in (text or '')


def best_text(*values):
    for value in values:
        text = compact_text(value)
        if text and not is_truncated_text(text):
            return text
    for value in values:
        text = compact_text(value)
        if text:
            return text
    return ''


def is_menu_text(text):
    return bool(MENU_TEXT_RE.search(text or ''))


def is_valid_result(item):
    name = compact_text(item.get('name', ''))
    if not name:
        return False
    price = item.get('price')
    if price is None:
        return False
    try:
        if float(price) <= 0:
            return False
    except (TypeError, ValueError):
        return False
    if is_menu_text(name):
        return False
    if GENERIC_NAME_RE.fullmatch(name):
        return False
    unit = compact_text(item.get('unit', ''))
    if re.fullmatch(r'\d+(?:\.\d+)?\s*(?:v|kv|โวลต์)', unit, flags=re.I):
        return False
    words = re.findall(r'[A-Za-zก-๙0-9]+', name)
    menu_words = re.findall(r'(Search|Product|Category|Home|Download|Contact|About|Shop)', name, flags=re.I)
    if words and len(menu_words) / max(len(words), 1) > 0.25:
        return False
    return True


def clean_result_name(name):
    name = compact_text(name)
    if len(name) > 180:
        name = name[:177].rstrip() + '...'
    return name


def matched_terms_for_text(text, q):
    hay = (text or '').lower()
    hay_norm = norm(text or '')
    out = []
    for term in expand_query_aliases(q):
        if term.lower() in hay or norm(term) in hay_norm:
            out.append(term)
    return out


def parse_html_tables(html, url, q, source_name, parser_name):
    soup = BeautifulSoup(html or '', 'html.parser')
    results = []
    current_group_name = ''
    for tr in soup.select('tr'):
        cells = [compact_text(c.get_text(' ', strip=True)) for c in tr.find_all(['td', 'th'])]
        cells = [c for c in cells if c]
        if not cells:
            continue
        row_text = ' '.join(cells)
        if is_menu_text(row_text):
            continue
        price_matches = [(i, PRICE_UNIT_RE.search(c)) for i, c in enumerate(cells)]
        price_matches = [(i, m) for i, m in price_matches if m]
        if not price_matches:
            if len(cells) <= 3 and len(row_text) <= 140 and matched_terms_for_text(row_text, q):
                current_group_name = row_text
            continue
        price_idx, price_match = price_matches[0]
        price = parse_price(price_match.group(1))
        unit = price_match.group(2) or ''
        if price is None or price <= 0:
            continue
        candidates = []
        if price_idx > 0:
            candidates.append(cells[price_idx - 1])
        candidates.extend(c for i, c in enumerate(cells) if i != price_idx and not PRICE_UNIT_RE.search(c))
        name = best_text(*candidates)
        name = re.sub(r'ราคา\s*[\d,]+(?:\.\d+)?\s*บาท', '', name, flags=re.I)
        name = PRICE_UNIT_RE.sub('', name)
        name = compact_text(name)
        if current_group_name and (len(name) < 10 or not matched_terms_for_text(name, q)):
            name = compact_text(current_group_name + ' ' + name)
        if not name or is_menu_text(name):
            continue
        terms = matched_terms_for_text(name + ' ' + row_text + ' ' + url, q)
        results.append(make_result(name, price, unit, source=source_name, url=url,
                                   parser_name=parser_name, matched_terms=terms, confidence=0.72))
    return results


def _json_ld_nodes(obj):
    if isinstance(obj, list):
        for item in obj:
            yield from _json_ld_nodes(item)
    elif isinstance(obj, dict):
        yield obj
        for key in ('@graph', 'hasVariant'):
            if key in obj:
                yield from _json_ld_nodes(obj[key])


def parse_json_ld_products(html, url, source_name):
    soup = BeautifulSoup(html or '', 'html.parser')
    out = []
    for sc in soup.find_all('script', type='application/ld+json'):
        try:
            data = json.loads(sc.string or '')
        except Exception:
            continue
        for node in _json_ld_nodes(data):
            typ = node.get('@type')
            types = typ if isinstance(typ, list) else [typ]
            if 'Product' not in types:
                continue
            offers = node.get('offers') or {}
            offers = offers if isinstance(offers, list) else [offers]
            for offer in offers:
                price = parse_price(str(offer.get('price', '') if isinstance(offer, dict) else ''))
                if price is None:
                    continue
                product_url = node.get('url') or (offer.get('url') if isinstance(offer, dict) else '') or url
                out.append(make_result(node.get('name', ''), price, source=source_name,
                                       url=absolutize_url(product_url, url), parser_name='json_ld',
                                       confidence=0.82))
    return out


def parse_woocommerce_products(base_url, q, source_name):
    base = source_origin(base_url).rstrip('/')
    r = fetch_url(base + '/wp-json/wc/store/v1/products',
                  params={'search': q, 'per_page': 15})
    out = []
    for p in r.json():
        prices = p.get('prices') or {}
        minor = prices.get('currency_minor_unit', 2)
        price = float(prices.get('price') or 0) / (10 ** minor)
        if price <= 0:
            continue
        name = BeautifulSoup(p.get('name', ''), 'html.parser').get_text(' ', strip=True)
        out.append(make_result(name, price, unit='', source=source_name,
                               url=p.get('permalink', base), parser_name='woocommerce',
                               confidence=0.9))
    return out


def parse_product_cards(html, url, source_name):
    soup = BeautifulSoup(html or '', 'html.parser')
    out = []
    selectors = 'li.product, div.product, .product-item, .product-card, a.productItem, [class*=product]'
    for card in soup.select(selectors):
        name_node = card.select_one('.woocommerce-loop-product__title, .product_name, h1, h2, h3, .title, .name, [class*=name], [class*=title]')
        price_node = card.select_one('.product_price, .price, .amount, [class*=price]')
        img = card.select_one('img[alt]')
        link = card if card.name == 'a' and card.get('href') else card.select_one('a[href]')
        price, unit = parse_price_unit(price_node.get_text(' ', strip=True) if price_node else card.get_text(' ', strip=True))
        if price is None or price <= 0:
            continue
        name = best_text(
            name_node.get('title') if name_node else '',
            card.get('title'),
            img.get('alt') if img else '',
            name_node.get_text(' ', strip=True) if name_node else '',
        )
        if not name:
            name = re.sub(r'[\d,]+(?:\.\d+)?\s*(?:บาท|/)', '', card.get_text(' ', strip=True)).strip()
        out.append(make_result(name, price, unit, source=source_name,
                               url=absolutize_url(link.get('href'), url) if link else url,
                               parser_name='product_card', confidence=0.68))
    return out


def parse_sirichai_product_cards(html, url, q):
    soup = BeautifulSoup(html or '', 'html.parser')
    out = []
    for card in soup.select('.so_sweet003'):
        text = compact_text(card.get_text(' ', strip=True))
        price, unit = parse_price_unit(text)
        if price is None or price <= 0:
            continue
        link = card.select_one('a[href]')
        name_node = card.select_one('span[style*="09429A"]')
        name = name_node.get_text(' ', strip=True) if name_node else ''
        if not name:
            name = re.split(r'\s+(?:Price|Model)\s*:?', text, 1, flags=re.I)[0]
        terms = matched_terms_for_text(name + ' ' + text + ' ' + (link.get('href', '') if link else ''), q)
        out.append(make_result(name, price, unit, source='sirichaielectric.com',
                               url=absolutize_url(link.get('href'), url) if link else url,
                               parser_name='sirichai_product_card',
                               matched_terms=terms, confidence=0.86))
    return out


def parse_sirichai_price_blocks(html, url, q):
    """Sirichai product pages render price tables as div pairs inside a table cell."""
    soup = BeautifulSoup(html or '', 'html.parser')
    out = []

    def append_row(name, price_text, row_url, confidence=0.9):
        price, unit = parse_price_unit(price_text)
        if price is None or price <= 0:
            return
        name = compact_text(name)
        if not name or is_menu_text(name):
            return
        terms = matched_terms_for_text(name + ' ' + row_url, q)
        out.append(make_result(name, price, unit, source='sirichaielectric.com',
                               url=absolutize_url(row_url, url),
                               parser_name='sirichai_price_table',
                               matched_terms=terms, confidence=confidence))

    # Price list rows: product link cell followed by price/unit cell.
    for a in soup.select('a[href*="/product_in/"]'):
        name = compact_text(a.get_text(' ', strip=True))
        if not name or is_menu_text(name):
            continue
        row_containers = []
        node = a
        for _ in range(5):
            node = node.parent
            if not node:
                break
            row_containers.append(node)
        for row in row_containers:
            found = False
            for sibling in row.find_next_siblings():
                sib_text = compact_text(sibling.get_text(' ', strip=True))
                if not sib_text:
                    continue
                if PRICE_UNIT_RE.search(sib_text):
                    append_row(name, sib_text, a.get('href', url), 0.92)
                    found = True
                    break
                if compact_text(sibling.get_text(' ', strip=True)) and len(sib_text) > 120:
                    break
            if found:
                break
    return out


def strict_filter_results(results, q):
    aliases = expand_query_aliases(q)
    technical = any(a.upper() in ('IEC', 'THW', 'YAZAKI') or a in ('60227', '60227 IEC', '60227 IEC 01', 'ยาซากิ')
                    for a in aliases)
    if not technical:
        return results
    filtered = []
    for item in results:
        hay = str(item.get('name', '') or '').lower()
        hay_norm = norm(hay)
        matches = [a for a in aliases if a.lower() in hay or norm(a) in hay_norm]
        if matches:
            item['matched_terms'] = sorted(set((item.get('matched_terms') or []) + matches))
            item['confidence'] = max(float(item.get('confidence') or 0), 0.75)
            filtered.append(item)
    return filtered


def dedupe_results(results):
    def priority(item):
        parser = item.get('parser_name', '')
        parser_score = {
            'sirichai_price_table': 50,
            'sirichai_search_table': 45,
            'sirichai_link_table': 45,
            'woocommerce': 40,
            'json_ld': 35,
            'sirichai_product_card': 30,
            'product_card': 20,
        }.get(parser, 10)
        return (parser_score, float(item.get('confidence') or 0), 1 if item.get('unit') else 0)

    keyed, order = {}, []
    for item in results:
        key = (
            norm(item.get('name', '')),
            item.get('source', ''),
            round(float(item.get('price') or 0), 2),
            item.get('unit', ''),
        )
        url_key = item.get('url') or ''
        dedupe_key = url_key if url_key else repr(key)
        if dedupe_key not in keyed:
            keyed[dedupe_key] = item
            order.append(dedupe_key)
        elif priority(item) > priority(keyed[dedupe_key]):
            keyed[dedupe_key] = item
    return [keyed[k] for k in order]


def normalize_results(results, q):
    normalized = []
    for item in results:
        if not item or not is_valid_result(item):
            continue
        item['name'] = clean_result_name(item.get('name', ''))
        terms = item.get('matched_terms') or matched_terms_for_text(
            '%s %s %s' % (item.get('name', ''), item.get('url', ''), item.get('source', '')), q)
        normalized_item = make_result(
            item.get('name', ''), item.get('price'), item.get('unit', ''),
            item.get('labor'), item.get('source', ''), item.get('url', ''),
            item.get('parser_name', ''), terms, item.get('confidence', 0.0)
        )
        if is_valid_result(normalized_item):
            normalized.append(normalized_item)
    normalized = strict_filter_results(normalized, q)
    normalized = dedupe_results(normalized)
    normalized.sort(key=lambda x: (x['price'] is None, x['price'] or 0, x['name']))
    return normalized


# ---------------------------------------------------------------- source-specific adapters
def search_sirichai(q):
    """sirichaielectric.com: search page + relevant cable/table pages."""
    source = 'sirichaielectric.com'
    base = 'https://sirichaielectric.com/'
    search_url = base + 'search_product.php?search=' + urllib.parse.quote(q)
    pages = []
    results = []
    r = fetch_url(base + 'search_product.php', params={'search': q})
    pages.append((r.text, r.url or search_url))
    results.extend(parse_html_tables(r.text, r.url or search_url, q, source, 'sirichai_search_table'))
    results.extend(parse_sirichai_price_blocks(r.text, r.url or search_url, q))
    results.extend(parse_sirichai_product_cards(r.text, r.url or search_url, q))
    results.extend(parse_product_cards(r.text, r.url or search_url, source))

    soup = BeautifulSoup(r.text, 'html.parser')
    aliases = expand_query_aliases(q)
    link_scores = {}
    for a in soup.find_all('a', href=True):
        text = compact_text(a.get_text(' ', strip=True))
        href = a['href']
        hay = (text + ' ' + href).lower()
        hay_norm = norm(hay)
        if any(alias.lower() in hay or norm(alias) in hay_norm for alias in aliases):
            full = absolutize_url(href, base)
            score = 0
            for alias in aliases:
                if alias.lower() in hay or norm(alias) in hay_norm:
                    score += 2
            if 'product_in' in hay:
                score += 2
            if any(token in hay for token in ('60227', 'iec-01', 'iec 01', 'thw', 'yazaki', 'ยาซากิ')):
                score += 4
            link_scores[full] = max(link_scores.get(full, 0), score)

    link_urls = [u for u, _ in sorted(link_scores.items(), key=lambda item: (-item[1], item[0]))]
    for full in link_urls[:24]:
        try:
            pr = fetch_url(full)
            results.extend(parse_html_tables(pr.text, pr.url or full, q, source, 'sirichai_link_table'))
            results.extend(parse_sirichai_price_blocks(pr.text, pr.url or full, q))
            results.extend(parse_sirichai_product_cards(pr.text, pr.url or full, q))
            results.extend(parse_json_ld_products(pr.text, pr.url or full, source))
            results.extend(parse_product_cards(pr.text, pr.url or full, source))
        except Exception:
            continue

    return normalize_results(results, q)


def search_ranfaifa(q):
    """ranfaifa.com (LnwShop): AJAX POST /search/box"""
    h = dict(UA)
    h.update({'X-Requested-With': 'XMLHttpRequest', 'Referer': 'https://www.ranfaifa.com/'})
    r = fetch_url('https://www.ranfaifa.com/search/box', method='POST', data={'q': q}, headers=h)
    soup = BeautifulSoup(r.text, 'html.parser')
    out = []
    for a in soup.select('a.productItem'):
        nm = a.select_one('.product_name')
        pz = a.select_one('.product_price')
        price = parse_price(pz.get_text()) if pz else None
        if nm and price:
            img = a.select_one('img[alt]')
            name = best_text(
                nm.get('title'),
                a.get('title'),
                img.get('alt') if img else '',
                nm.get_text(' ', strip=True),
            )
            url = absolutize_url(a.get('href', ''), 'https://www.ranfaifa.com/')
            out.append(make_result(name, price, source='ranfaifa.com', url=url,
                                   parser_name='ranfaifa_ajax', confidence=0.72))

    def fetch_full_name(item):
        if not item.get('url'):
            return item
        try:
            pr = fetch_url(item['url'])
            ps = BeautifulSoup(pr.text, 'html.parser')
            og = ps.select_one('meta[property="og:title"], meta[name="twitter:title"]')
            title = ps.select_one('title')
            candidates = []
            for sel in ('h1', '.product_name', '.productName', '.product-title', '.product_title'):
                candidates.extend(node.get_text(' ', strip=True) for node in ps.select(sel))
            candidates.extend([
                og.get('content') if og else '',
                title.get_text(' ', strip=True) if title else '',
            ])
            full = best_text(*candidates)
            if full:
                full = re.split(r'\s*[-|]\s*ร้าน|\s*[-|]\s*Ranfaifa|\s*[-|]\s*รันไฟฟ้า', full, 1, flags=re.I)[0].strip()
            if full and (is_truncated_text(item['name']) or len(full) > len(item['name'])):
                item['name'] = full
        except Exception:
            pass
        return item

    with ThreadPoolExecutor(max_workers=5) as ex:
        enriched = []
        for f in as_completed([ex.submit(fetch_full_name, item) for item in out[:12]]):
            enriched.append(f.result())
        if len(out) > 12:
            enriched.extend(out[12:])
        return normalize_results(enriched, q)


def search_obec(q):
    """บัญชีราคาวัสดุ+ค่าแรง สพฐ. ปี 2568 (extract จาก PDF ไว้ล่วงหน้า)"""
    if not os.path.exists(OBEC_FILE):
        return []
    with open(OBEC_FILE, encoding='utf-8') as f:
        data = json.load(f)
    qn = norm(q)
    words = [norm(w) for w in q.split() if norm(w)]
    aliases = [norm(a) for a in expand_query_aliases(q)]
    out = []
    for it in data['items']:
        searchable = norm(' '.join([it.get('name', ''), it.get('code', ''), it.get('note', '')]))
        if qn in searchable or (words and all(w in searchable for w in words)) or any(a and a in searchable for a in aliases):
            name = it['name'] + (' [%s]' % it['code'] if it['code'] else '')
            if it.get('note'):
                name += ' (%s)' % it['note']
            out.append(make_result(name, it['material'], it['unit'], it['labor'],
                                   'YOTATHAI / OBEC 2568', data['source_url'],
                                   'obec_json', matched_terms=matched_terms_for_text(searchable, q),
                                   confidence=0.8))
        if len(out) >= 30:
            break
    return normalize_results(out, q)


def search_apelectric(q):
    """apelectricstore.com: WooCommerce Store API (ราคา 0 = ต้องสอบถามร้าน -> ข้าม)"""
    return normalize_results(parse_woocommerce_products('https://apelectricstore.com', q, 'apelectricstore.com'), q)


def search_custom(q, url):
    """แหล่งที่ user เพิ่มเอง: ลอง WooCommerce API -> JSON-LD -> CSS selector ทั่วไป"""
    base = source_origin(url).rstrip('/')
    source = urllib.parse.urlparse(base).netloc
    out = []
    # 1) WooCommerce Store API
    try:
        out.extend(parse_woocommerce_products(base, q, source))
    except Exception:
        pass
    # 2) หน้า search ทั่วไป: รองรับ URL base, URL ที่มี query เดิม, และ template {q}
    try:
        search_url = build_custom_search_url(url, q)
        r = fetch_url(search_url)
        out.extend(parse_json_ld_products(r.text, r.url or search_url, source))
        out.extend(parse_html_tables(r.text, r.url or search_url, q, source, 'custom_table'))
        out.extend(parse_product_cards(r.text, r.url or search_url, source))
    except Exception:
        pass
    return normalize_results(out, q)


SOURCES = {
    'sirichai': search_sirichai,
    'ranfaifa': search_ranfaifa,
    'obec': search_obec,
    'apelectric': search_apelectric,
}


# ---------------------------------------------------------------- Flask routes
@app.route('/api/search')
def api_search():
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify({'results': [], 'errors': [], 'source_status': {}})
    wanted = request.args.get('sources', ','.join(SOURCES)).split(',')
    custom_urls = [u for u in request.args.get('custom', '').split('|') if u.strip()]
    source_status = {
        'sirichai': make_source_status('no_result', 0, 'ไม่ได้เลือกแหล่งนี้'),
        'ranfaifa': make_source_status('no_result', 0, 'ไม่ได้เลือกแหล่งนี้'),
        'obec': make_source_status('no_result', 0, 'ไม่ได้เลือกแหล่งนี้'),
        'apelectric': make_source_status('no_result', 0, 'ไม่ได้เลือกแหล่งนี้'),
        'custom': make_source_status('no_result', 0, 'ไม่ได้เพิ่ม Custom URL'),
    }
    results, errors = [], []
    with ThreadPoolExecutor(max_workers=6) as ex:
        futs = {}
        for key in wanted:
            if key in SOURCES:
                futs[ex.submit(SOURCES[key], q)] = key
        for cu in custom_urls:
            futs[ex.submit(search_custom, q, cu.strip())] = 'custom'
        for f in as_completed(futs):
            source_key = futs[f]
            try:
                found = f.result()
                results.extend(found)
                if source_key == 'custom':
                    old = source_status['custom']
                    count = old['result_count'] + len(found)
                    message = '' if count else 'ไม่สามารถดึงราคาอัตโนมัติจากเว็บนี้ได้ กรุณาเปิดเว็บต้นทางและเพิ่มราคาด้วยตนเอง'
                    source_status['custom'] = make_source_status('success' if count else 'no_result', count, message)
                else:
                    source_status[source_key] = make_source_status(
                        'success' if found else 'no_result',
                        len(found),
                        '' if found else 'ไม่พบราคาที่ตรงกับคำค้น'
                    )
            except Exception as e:
                errors.append('%s: %s' % (source_key, e))
                if source_key in source_status:
                    source_status[source_key] = make_source_status('error', 0, str(e))
    for r in results:
        r['fetched_at'] = today()
    results = dedupe_results(results)
    results.sort(key=lambda x: (x['price'] is None, x['price'] or 0, x.get('name', '')))
    return jsonify({'results': results, 'errors': errors, 'source_status': source_status})


# ---------------------------------------------------------------- library
def load_lib():
    if os.path.exists(LIB_FILE):
        with open(LIB_FILE, encoding='utf-8') as f:
            return json.load(f)
    return []


def save_lib(lib):
    with open(LIB_FILE, 'w', encoding='utf-8') as f:
        json.dump(lib, f, ensure_ascii=False, indent=1)


@app.route('/api/library', methods=['GET', 'POST', 'DELETE'])
def api_library():
    with _lib_lock:
        lib = load_lib()
        if request.method == 'GET':
            return jsonify(lib)
        if request.method == 'POST':
            item = request.get_json(force=True)
            item['id'] = (max((i['id'] for i in lib), default=0) + 1)
            item['saved_at'] = today()
            lib.append(item)
            save_lib(lib)
            return jsonify(item)
        # DELETE ?id=
        did = int(request.args.get('id', -1))
        lib = [i for i in lib if i['id'] != did]
        save_lib(lib)
        return jsonify({'ok': True})


@app.route('/api/library/update', methods=['POST'])
def api_library_update():
    patch = request.get_json(force=True)
    with _lib_lock:
        lib = load_lib()
        for i in lib:
            if i['id'] == patch['id']:
                i.update(patch)
        save_lib(lib)
    return jsonify({'ok': True})


@app.route('/api/library/export')
def api_library_export():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    lib = load_lib()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Price Library'
    heads = ['#', 'รายการ', 'ราคาฐาน (บาท)', 'Markup %', 'ราคาใช้จริง (บาท)',
             'หน่วย', 'ปริมาณ', 'รวมเงิน (บาท)', 'ค่าแรง (บาท)',
             'แหล่งที่มา', 'URL', 'วันที่ดึงข้อมูล', 'วันที่บันทึก']
    ws.append(heads)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='305496')
        c.alignment = Alignment(horizontal='center')
    grand = 0.0
    for n, it in enumerate(lib, 1):
        qty = it.get('qty', 1) or 0
        fp = it.get('final_price')
        total = round(fp * qty, 2) if fp is not None else None
        grand += total or 0
        ws.append([n, it.get('name'), it.get('price'), it.get('markup_pct'),
                   fp, it.get('unit'), qty, total, it.get('labor'),
                   it.get('source'), it.get('url'), it.get('fetched_at'), it.get('saved_at')])
    ws.append([None, 'รวมทั้งหมด', None, None, None, None, None, round(grand, 2)])
    last = ws[ws.max_row]
    last[1].font = Font(bold=True)
    last[7].font = Font(bold=True)
    widths = [5, 55, 14, 10, 16, 10, 10, 15, 12, 28, 45, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    out = os.path.join(DATA_DIR, 'price_library_%s.xlsx' % today())
    wb.save(out)
    return send_file(out, as_attachment=True,
                     download_name=os.path.basename(out))


# ---------------------------------------------------------------- BOQ
def list_boq_files():
    return [f for f in os.listdir(BASE)
            if f.lower().endswith('.xlsx') and not f.startswith('~$')
            and 'price_library' not in f]


@app.route('/api/boq/files')
def api_boq_files():
    return jsonify(list_boq_files())


@app.route('/api/boq/items')
def api_boq_items():
    import openpyxl
    fname = request.args.get('file') or (list_boq_files() or [None])[0]
    if not fname:
        return jsonify({'error': 'ไม่พบไฟล์ .xlsx ในโฟลเดอร์'}), 404
    path = os.path.join(BASE, fname)
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = None
    for name in wb.sheetnames:
        if name.strip().lower() in ('detail', 'details'):
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb[wb.sheetnames[-1]]
    items, section = [], ''
    for r in range(1, sheet.max_row + 1):
        a = sheet.cell(r, 1).value
        b = sheet.cell(r, 2).value
        c = sheet.cell(r, 3).value
        d = sheet.cell(r, 4).value
        e = sheet.cell(r, 5).value
        g = sheet.cell(r, 7).value
        if a is not None and b and not isinstance(b, (int, float)):
            section = '%s. %s' % (a, b)
        if b and isinstance(c, (int, float)) and d:
            if str(b).strip().upper().startswith('SUM'):
                continue
            items.append({'row': r, 'section': section, 'name': str(b).strip(),
                          'qty': c, 'unit': str(d).strip(),
                          'material': e, 'labor': g})
    return jsonify({'file': fname, 'sheet': sheet.title, 'items': items})


@app.route('/api/boq/fill', methods=['POST'])
def api_boq_fill():
    import openpyxl
    body = request.get_json(force=True)
    fname = body['file']
    assigns = body['assignments']  # [{row, material, labor}]
    path = os.path.join(BASE, fname)
    wb = openpyxl.load_workbook(path)  # เก็บสูตรเดิมไว้
    sheet = None
    for name in wb.sheetnames:
        if name.strip().lower() in ('detail', 'details'):
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb[wb.sheetnames[-1]]
    n = 0
    for a in assigns:
        r = int(a['row'])
        if a.get('material') is not None:
            sheet.cell(r, 5).value = float(a['material'])   # E ค่าวัสดุ/หน่วย
            n += 1
        if a.get('labor') is not None:
            sheet.cell(r, 7).value = float(a['labor'])       # G ค่าแรง/หน่วย
    stem = re.sub(r'\.xlsx$', '', fname, flags=re.I)
    outname = '%s (filled %s).xlsx' % (stem, datetime.datetime.now().strftime('%Y-%m-%d %H%M'))
    wb.save(os.path.join(BASE, outname))
    return jsonify({'ok': True, 'filled': n, 'output': outname})


@app.route('/api/boq/download')
def api_boq_download():
    fname = request.args.get('file', '')
    if fname not in os.listdir(BASE):
        return jsonify({'error': 'not found'}), 404
    return send_file(os.path.join(BASE, fname), as_attachment=True, download_name=fname)


@app.route('/')
def index():
    return send_from_directory(BASE, 'index.html')


@app.route('/single')
def single_file_app():
    return send_from_directory(BASE, 'BOQ_price_finder.html')


@app.route('/BOQ_price_finder.html')
def single_file_app_html():
    return send_from_directory(BASE, 'BOQ_price_finder.html')


if __name__ == '__main__':
    # HOST=0.0.0.0 python3 server.py  -> เปิดให้เครื่องอื่นใน Wi-Fi วงเดียวกันเข้าได้
    host = os.environ.get('HOST', '0.0.0.0')
    port = int(os.environ.get('PORT', 5544))
    display_host = '127.0.0.1' if host in ('0.0.0.0', '127.0.0.1') else host
    print('BOQ Price Finder -> http://%s:%s' % (display_host, port))
    app.run(host=host, port=port, debug=False)
